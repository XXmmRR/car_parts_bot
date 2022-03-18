"""Тут будет вся логика связанная с бд"""
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, Integer, String, Float

DATABASE_NAME = 'bot.sqlite'

engine = create_engine(f"sqlite:///{DATABASE_NAME}")
Session = sessionmaker(bind=engine)

Base = declarative_base()


class BaseCars(Base):
    __tablename__ = 'cars'
    element_id = Column(Integer, primary_key=True)
    id = Column(String(10), nullable=True)
    mark = Column(String(150), nullable=True,)
    model = Column(String(150), nullable=True)
    generation = Column(String(150), nullable=True)
    year_from = Column(Integer, nullable=True)
    year_to = Column(Integer, nullable=True)
    body_type = Column(String(150), nullable=True)
    notice = Column(String(150), nullable=True)
    volume = Column(Integer, nullable=True)
    transmission = Column(String(150), nullable=True)
    complectation = Column(String(150), nullable=True)
    horse_power = Column(Integer, nullable=True)
    engine_type = Column(String(150), nullable=True)
    photo = Column(Integer, nullable=True)
    logo = Column(String(20), nullable=True)
    cyrillic_mark = Column(String(150), nullable=True,)
    cyrillic_model = Column(String(150), nullable=True)
    door_count = Column(Integer, nullable=True)
    seats = Column(String(50), nullable=True)
    length = Column(Integer, nullable=True)
    width = Column(Integer, nullable=True)
    height = Column(Integer, nullable=True)
    wheel_base = Column(Integer, nullable=True)
    clearance = Column(String(150), nullable=True)
    front_wheel_base = Column(Integer, nullable=True)
    back_wheel_base = Column(Integer, nullable=True)
    wheel_size = Column(String(150), nullable=True)
    trunks_min_capacity = Column(Integer, nullable=True)
    trunks_max_capacity = Column(Integer, nullable=True)
    weight = Column(Integer, nullable=True)
    full_weight = Column(Integer, nullable=True)
    gear_value = Column(Integer, nullable=True)
    gear_type = Column(String(150), nullable=True)
    front_suspension = Column(String(150), nullable=True)
    back_suspension = Column(String(150), nullable=True)
    front_brake = Column(String(150), nullable=True)
    back_brake = Column(String(150), nullable=True)
    max_speed = Column(Integer, nullable=True)
    time_to_100 = Column(Float, nullable=True)
    consumption_mixed = Column(Float, nullable=True)
    consumption_hiway = Column(Float, nullable=True)
    consumption_city = Column(Float, nullable=True)
    petrol_type = Column(String(150), nullable=True)
    emission_euro_class = Column(String(150), nullable=True)
    fuel_emission = Column(Integer, nullable=True)
    engine_order = Column(String(150), nullable=True)
    feeding = Column(String(150), nullable=True)
    cylinders_order = Column(String(150), nullable=True)
    cylinders_value = Column(Integer, nullable=True)
    valves = Column(Integer, nullable=True)
    engine_feeding = Column(String(150), nullable=True)
    compression = Column(Float, nullable=True)
    diametr = Column(Integer, nullable=True)
    piston_stroke = Column(Integer, nullable=True)
    moment = Column(Integer, nullable=True)
    moment_rpm = Column(String(50), nullable=True)
    kvt_power = Column(Integer, nullable=True)
    rpm_power = Column(String(150), nullable=True)
    fuel_tank_capacity = Column(Integer, nullable=True)


def create_db():
    Base.metadata.create_all(engine)
    session = Session()
    session.commit()



def load_data():
    from openpyxl import load_workbook
    from db import BaseCars, Session

    book = load_workbook('base_cars_full.xlsx', read_only=True)
    sheet = book.active

    rows = sheet.iter_rows()

    def load_data(rows):
        session = Session()
        for row in rows:
            cars = list(row[x].value for x in range(0, 59))
            print(len(cars))
            add_cars = BaseCars(id=cars[0], mark=cars[1], model=cars[2], generation=cars[3], year_from=cars[4],
                                year_to=cars[5], body_type=cars[6],
                                notice=cars[7], volume=cars[8], transmission=cars[9], complectation=cars[10],
                                horse_power=cars[11], engine_type=cars[12], photo=cars[13], logo=cars[14],
                                cyrillic_mark=cars[15],
                                cyrillic_model=cars[16], door_count=cars[17], seats=cars[18], length=cars[19],
                                width=cars[20],
                                height=cars[21], wheel_base=cars[22], clearance=cars[23], front_wheel_base=cars[24],
                                back_wheel_base=cars[25], wheel_size=cars[26], trunks_min_capacity=cars[27],
                                trunks_max_capacity=cars[28], weight=cars[29], full_weight=cars[30],
                                gear_value=cars[31],
                                gear_type=cars[32], front_suspension=cars[33], back_suspension=cars[34],
                                front_brake=cars[35],
                                back_brake=cars[36], max_speed=cars[37], time_to_100=cars[38],
                                consumption_mixed=cars[39],
                                consumption_hiway=cars[40], consumption_city=cars[41], petrol_type=cars[42],
                                emission_euro_class=cars[43], fuel_emission=cars[44], engine_order=cars[45],
                                feeding=cars[46], cylinders_order=cars[47], cylinders_value=cars[48], valves=cars[49],
                                engine_feeding=cars[50], compression=cars[51], diametr=cars[52], piston_stroke=cars[53],
                                moment=cars[54], moment_rpm=cars[55], kvt_power=cars[56],
                                rpm_power=cars[57], fuel_tank_capacity=cars[58], )

            session.add(add_cars)
            session.commit()
        session.close()


if __name__ == '__main__':
    create_db()