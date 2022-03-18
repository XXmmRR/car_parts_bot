from openpyxl import load_workbook
from db import BaseCars, Session

book = load_workbook('base_cars_full.xlsx', read_only=True)
sheet = book.active

rows = sheet.iter_rows()


def load_data(rows):
    session = Session()
    for row in rows:
        cars = list(row[x].value for x in range(0, 59))
        print(len(cars))
        add_cars = BaseCars(id=cars[0], mark=cars[1], model=cars[2], generation=cars[3], year_from=cars[4], year_to=cars[5], body_type=cars[6],
                            notice = cars[7], volume = cars[8], transmission = cars[9], complectation=cars[10],
                            horse_power=cars[11], engine_type=cars[12], photo=cars[13], logo=cars[14], cyrillic_mark=cars[15],
                            cyrillic_model=cars[16], door_count=cars[17], seats=cars[18], length=cars[19], width=cars[20],
                            height=cars[21], wheel_base=cars[22], clearance=cars[23], front_wheel_base=cars[24],
                            back_wheel_base=cars[25], wheel_size=cars[26], trunks_min_capacity=cars[27],
                            trunks_max_capacity=cars[28], weight=cars[29], full_weight=cars[30], gear_value=cars[31],
                            gear_type=cars[32], front_suspension=cars[33], back_suspension=cars[34], front_brake=cars[35],
                            back_brake=cars[36], max_speed=cars[37], time_to_100=cars[38], consumption_mixed=cars[39],
                            consumption_hiway=cars[40], consumption_city=cars[41], petrol_type=cars[42],
                            emission_euro_class=cars[43], fuel_emission=cars[44], engine_order=cars[45],
                            feeding=cars[46], cylinders_order=cars[47], cylinders_value=cars[48], valves=cars[49],
                            engine_feeding=cars[50], compression=cars[51], diametr=cars[52], piston_stroke=cars[53],
                            moment=cars[54], moment_rpm=cars[55], kvt_power=cars[56],
                            rpm_power=cars[57], fuel_tank_capacity=cars[58],)

        session.add(add_cars)
        session.commit()
    session.close()